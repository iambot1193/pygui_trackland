#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import os
import re
import unicodedata
from copy import copy
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ====== CAMINHOS FIXOS (os que você mostrou) ======
ENTRADA = r"C:\Users\jenni\Downloads\entrada_xlsx"
SAIDA   = r"C:\Users\jenni\Downloads\saida_xlsx"

# Cabeçalhos aceitos para a coluna de unidade (case/acento/espacos ignorados).
UNIT_HEADERS = [
    "Unidade organizacional cliente",
    "Unidade organizacional",
    "Unidade organizacional do cliente",
]

# Dica para reconhecer a linha do cabeçalho (deixe vazio se não quiser exigir nada).
REQUIRED_HINTS = ["Placa"]


def normalize(text) -> str:
    """Normaliza texto para comparar cabeçalhos (lower, sem acento, espaços compactados)."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s


UNIT_HEADERS_N = {normalize(h) for h in UNIT_HEADERS}
REQUIRED_HINTS_N = {normalize(h) for h in REQUIRED_HINTS}


def sanitize_filename(name: str, max_len: int = 150) -> str:
    """Gera um nome de arquivo seguro (sem caracteres inválidos, sem acento)."""
    if name is None or str(name).strip() == "":
        name = "SEM_UNIDADE"
    name = str(name).strip()

    s = unicodedata.normalize("NFKD", name)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))

    # Remove caracteres inválidos (Windows + controle)
    s = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", s)
    s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")

    if len(s) > max_len:
        s = s[:max_len].rstrip("_")
    return s or "SEM_UNIDADE"


def copy_cell(src, dst) -> None:
    """Copia valor + estilo (o suficiente para a maioria dos relatórios)."""
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)
    if src.comment:
        dst.comment = copy(src.comment)
    dst.hyperlink = src.hyperlink


def last_used_col(header_vals: List) -> int:
    """Última coluna realmente usada no cabeçalho (célula não vazia)."""
    last = 1
    for i, v in enumerate(header_vals, start=1):
        if v is not None and str(v).strip() != "":
            last = i
    return last


def find_header_and_unit_col(ws, max_scan_rows: int = 50) -> Tuple[int, int]:
    """
    Encontra:
      - linha do cabeçalho
      - coluna da unidade organizacional
    Procura nas primeiras `max_scan_rows` linhas.
    """
    max_row = min(max_scan_rows, ws.max_row or 1)
    max_col = ws.max_column or 1

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        row_n = [normalize(v) for v in row_vals]

        unit_positions = [i for i, v in enumerate(row_n, start=1) if v in UNIT_HEADERS_N]
        has_hints = REQUIRED_HINTS_N.issubset(set(row_n)) if REQUIRED_HINTS_N else True

        if unit_positions and has_hints:
            return r, unit_positions[0]

    # fallback: sem exigir hints
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        row_n = [normalize(v) for v in row_vals]
        unit_positions = [i for i, v in enumerate(row_n, start=1) if v in UNIT_HEADERS_N]
        if unit_positions:
            return r, unit_positions[0]

    raise ValueError(
        "Não encontrei a linha de cabeçalho com a coluna de Unidade Organizacional. "
        f"Cabeçalhos aceitos: {UNIT_HEADERS}"
    )


def copy_column_dimensions(src_ws, dst_ws, max_col: int) -> None:
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        if col_letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width


def split_one_file(input_path: str, out_root: str) -> List[str]:
    """
    Divide o primeiro sheet do arquivo por unidade organizacional.
    Retorna a lista de arquivos criados.
    """
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    header_row, unit_col = find_header_and_unit_col(ws)
    header_vals = [ws.cell(header_row, c).value for c in range(1, (ws.max_column or 1) + 1)]
    max_col = last_used_col(header_vals)

    # Agrupa linhas por unidade
    groups: Dict[str, List[int]] = {}
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        unit_val = ws.cell(r, unit_col).value
        unit_key = str(unit_val).strip() if unit_val is not None and str(unit_val).strip() else "SEM_UNIDADE"
        groups.setdefault(unit_key, []).append(r)

    base = os.path.splitext(os.path.basename(input_path))[0]
    out_dir = os.path.join(out_root, base)
    os.makedirs(out_dir, exist_ok=True)

    created: List[str] = []

    for unit, rows in groups.items():
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        out_ws = out_wb.create_sheet(title=(ws.title or "Dados")[:31])

        copy_column_dimensions(ws, out_ws, max_col)

        # Cabeçalho -> linha 1
        for c in range(1, max_col + 1):
            copy_cell(ws.cell(header_row, c), out_ws.cell(1, c))

        # Dados -> a partir da linha 2
        out_r = 2
        for src_r in rows:
            for c in range(1, max_col + 1):
                copy_cell(ws.cell(src_r, c), out_ws.cell(out_r, c))

            # Altura de linha (opcional)
            if src_r in ws.row_dimensions and ws.row_dimensions[src_r].height:
                out_ws.row_dimensions[out_r].height = ws.row_dimensions[src_r].height
            out_r += 1

        # Congelar painéis (header)
        out_ws.freeze_panes = "A2"

        # Auto filtro
        out_ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{out_ws.max_row}"

        # Merges do cabeçalho (se existirem)
        for mr in ws.merged_cells.ranges:
            if mr.min_row == header_row and mr.max_row == header_row and mr.max_col <= max_col:
                new_range = f"{get_column_letter(mr.min_col)}1:{get_column_letter(mr.max_col)}1"
                out_ws.merge_cells(new_range)

        fname = sanitize_filename(unit) + ".xlsx"
        out_path = os.path.join(out_dir, fname)

        # Evita sobrescrever se existir
        if os.path.exists(out_path):
            i = 2
            while True:
                alt = os.path.join(out_dir, sanitize_filename(unit) + f"_{i}.xlsx")
                if not os.path.exists(alt):
                    out_path = alt
                    break
                i += 1

        out_wb.save(out_path)
        created.append(out_path)

    return created


def iter_input_files(folder: str) -> List[str]:
    """Pega .xlsx/.xlsm diretamente na pasta ENTRADA (sem recursão)."""
    exts = {".xlsx", ".xlsm"}
    found: List[str] = []
    for fn in os.listdir(folder):
        fp = os.path.join(folder, fn)
        if os.path.isfile(fp) and os.path.splitext(fn)[1].lower() in exts:
            found.append(fp)

    if not found:
        raise FileNotFoundError("Não encontrei arquivos .xlsx/.xlsm na pasta de entrada.")
    return found


def main() -> None:
    os.makedirs(SAIDA, exist_ok=True)

    files = iter_input_files(ENTRADA)
    total_out = 0

    for f in files:
        created = split_one_file(f, SAIDA)
        total_out += len(created)
        print(f"[OK] {os.path.basename(f)} -> {len(created)} arquivos")

    print(f"Finalizado. Total de arquivos gerados: {total_out}")
    print(f"Saída em: {SAIDA}")


if __name__ == "__main__":
    main()
