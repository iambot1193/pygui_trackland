#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import os
import re
import unicodedata
from copy import copy
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ====== CAMINHOS FIXOS ======
ENTRADA = r"C:\Users\jenni\Downloads\entrada_xlsx"
SAIDA   = r"C:\Users\jenni\Downloads\saida_xlsx"

UNIT_HEADERS = [
    "Unidade organizacional cliente",
    "Unidade organizacional",
    "Unidade organizacional do cliente",
]

REQUIRED_HINTS = ["Placa"]


def normalize(text) -> str:
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s


UNIT_HEADERS_N = {normalize(h) for h in UNIT_HEADERS}
REQUIRED_HINTS_N = {normalize(h) for h in REQUIRED_HINTS}


def sanitize_filename(name: str, max_len: int = 150) -> str:
    if name is None or str(name).strip() == "":
        name = "SEM_UNIDADE"
    name = str(name).strip()

    s = unicodedata.normalize("NFKD", name)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))

    s = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", s)
    s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")

    if len(s) > max_len:
        s = s[:max_len].rstrip("_")
    return s or "SEM_UNIDADE"


def bold_font(f: Font | None) -> Font:
    """Força negrito preservando o resto (compatível com openpyxl antigo)."""
    if f is None:
        return Font(b=True)
    return Font(
        name=f.name,
        sz=f.sz,
        b=True,
        i=f.i,
        charset=f.charset,
        u=f.u,
        strike=f.strike,
        color=f.color,
        family=f.family,
        vertAlign=f.vertAlign,
        scheme=f.scheme,
        outline=f.outline,
        shadow=f.shadow,
        condense=f.condense,
        extend=f.extend
    )


def copy_cell(src, dst) -> None:
    """Copia valor + estilo e força a fonte em negrito."""
    dst.value = src.value

    if src.has_style:
        dst.font = bold_font(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)
    else:
        dst.font = bold_font(None)

    if src.comment:
        dst.comment = copy(src.comment)
    dst.hyperlink = src.hyperlink


def last_used_col(header_vals: List) -> int:
    last = 1
    for i, v in enumerate(header_vals, start=1):
        if v is not None and str(v).strip() != "":
            last = i
    return last


def find_header_and_unit_col(ws, max_scan_rows: int = 50) -> Tuple[int, int]:
    max_row = min(max_scan_rows, ws.max_row or 1)
    max_col = ws.max_column or 1

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        row_n = [normalize(v) for v in row_vals]

        unit_positions = [i for i, v in enumerate(row_n, start=1) if v in UNIT_HEADERS_N]
        has_hints = REQUIRED_HINTS_N.issubset(set(row_n)) if REQUIRED_HINTS_N else True

        if unit_positions and has_hints:
            return r, unit_positions[0]

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        row_n = [normalize(v) for v in row_vals]
        unit_positions = [i for i, v in enumerate(row_n, start=1) if v in UNIT_HEADERS_N]
        if unit_positions:
            return r, unit_positions[0]

    raise ValueError(
        "Não encontrei a linha de cabeçalho com a coluna de Unidade Organizacional. "
        f"Cabeçalhos aceitos: {UNIT_HEADERS}"
    )


def row_is_visually_blank(ws_values, r: int, max_col: int) -> bool:
    """Considera a linha vazia olhando data_only=True (resultado exibido)."""
    for c in range(1, max_col + 1):
        v = ws_values.cell(r, c).value
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def clear_freeze_and_splits(out_ws) -> None:
    """Remove congelamento/split (linha cinza no Sheets)."""
    out_ws.freeze_panes = None
    sv = out_ws.sheet_view
    try:
        sv.splitRow = 0
        sv.splitColumn = 0
        sv.pane = None
    except Exception:
        pass


# ====== NOVO: TAMANHOS PADRÃO ======
ALTURA_PADRAO_LINHA = 30.0
LARGURA_PADRAO_COLUNA = 26.0


def set_default_sizes(ws, max_col: int,
                      row_height: float = ALTURA_PADRAO_LINHA,
                      col_width: float = LARGURA_PADRAO_COLUNA) -> None:
    """Aplica altura/largura fixas para todas as linhas/colunas usadas."""
    # colunas
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_width

    # linhas
    for r in range(1, (ws.max_row or 1) + 1):
        ws.row_dimensions[r].height = row_height


def split_one_file(input_path: str, out_root: str) -> List[str]:
    wb = load_workbook(input_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    wb_vals = load_workbook(input_path, data_only=True)
    ws_vals = wb_vals[wb_vals.sheetnames[0]]

    header_row, unit_col = find_header_and_unit_col(ws)
    header_vals = [ws.cell(header_row, c).value for c in range(1, (ws.max_column or 1) + 1)]
    max_col = last_used_col(header_vals)

    groups: Dict[str, List[int]] = {}
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if row_is_visually_blank(ws_vals, r, max_col):
            continue

        unit_val = ws.cell(r, unit_col).value
        unit_key = str(unit_val).strip() if unit_val is not None and str(unit_val).strip() else "SEM_UNIDADE"
        groups.setdefault(unit_key, []).append(r)

    base = os.path.splitext(os.path.basename(input_path))[0]
    out_dir = os.path.join(out_root, base)
    os.makedirs(out_dir, exist_ok=True)

    created: List[str] = []

    for unit, rows in groups.items():
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        out_ws = out_wb.create_sheet(title=(ws.title or "Dados")[:31])

        # Cabeçalho
        for c in range(1, max_col + 1):
            copy_cell(ws.cell(header_row, c), out_ws.cell(1, c))

        # Dados
        out_r = 2
        for src_r in rows:
            for c in range(1, max_col + 1):
                copy_cell(ws.cell(src_r, c), out_ws.cell(out_r, c))
            out_r += 1

        last_row = out_ws.max_row or 1

        # filtro
        out_ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{last_row}"

        # merges do cabeçalho
        for mr in ws.merged_cells.ranges:
            if mr.min_row == header_row and mr.max_row == header_row and mr.max_col <= max_col:
                new_range = f"{get_column_letter(mr.min_col)}1:{get_column_letter(mr.max_col)}1"
                out_ws.merge_cells(new_range)

        # remove freeze/split (linha cinza)
        clear_freeze_and_splits(out_ws)

        # ✅ aplica tamanho fixo (altura e largura)
        set_default_sizes(out_ws, max_col=max_col,
                          row_height=ALTURA_PADRAO_LINHA,
                          col_width=LARGURA_PADRAO_COLUNA)

        fname = sanitize_filename(unit) + ".xlsx"
        out_path = os.path.join(out_dir, fname)

        if os.path.exists(out_path):
            i = 2
            while True:
                alt = os.path.join(out_dir, sanitize_filename(unit) + f"_{i}.xlsx")
                if not os.path.exists(alt):
                    out_path = alt
                    break
                i += 1

        out_wb.save(out_path)
        created.append(out_path)

    return created


def iter_input_files(folder: str) -> List[str]:
    exts = {".xlsx", ".xlsm"}
    found: List[str] = []
    for fn in os.listdir(folder):
        fp = os.path.join(folder, fn)
        if os.path.isfile(fp) and os.path.splitext(fn)[1].lower() in exts:
            found.append(fp)
    if not found:
        raise FileNotFoundError("Não encontrei arquivos .xlsx/.xlsm na pasta de entrada.")
    return found


def main() -> None:
    os.makedirs(SAIDA, exist_ok=True)

    files = iter_input_files(ENTRADA)
    total_out = 0

    for f in files:
        created = split_one_file(f, SAIDA)
        total_out += len(created)
        print(f"[OK] {os.path.basename(f)} -> {len(created)} arquivos")

    print(f"Finalizado. Total de arquivos gerados: {total_out}")
    print(f"Saída em: {SAIDA}")


if __name__ == "__main__":
    main()
