import os
import unicodedata
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date
from pathlib import Path

# ===============================================================
# FUNÇÃO: normalizar texto (tirar acentos, ç, etc.)
# ===============================================================
def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    # Remove acentos
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    # Garante Ç -> C (caso ainda reste)
    texto = texto.replace("Ç", "C")
    return texto

# ===============================================================
# FUNÇÃO: copia células preservando TUDO (cores, fonte, bordas)
# ===============================================================
def copiar_celula(origem, destino):
    destino.value = origem.value

    if origem.has_style:
        destino.font = origem.font.copy()
        destino.border = origem.border.copy()
        destino.fill = origem.fill.copy()
        destino.number_format = origem.number_format
        destino.protection = origem.protection.copy()
        destino.alignment = origem.alignment.copy()

# ===============================================================
# FUNÇÃO: formatar o arquivo final (como você já fazia)
# ===============================================================
def formatar_arquivo_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    fonte = Font(name="Calibri", size=12, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    fill_cabecalho = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    max_row = ws.max_row
    max_col = 10

    # cabeçalho
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        # só pinta se for sem cor / branco
        if cell.fill.start_color.index in ("00000000", "FFFFFFFF"):
            cell.fill = fill_cabecalho
        cell.font = fonte
        cell.alignment = alinhamento
        cell.border = borda

    # corpo
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = fonte
            cell.alignment = alinhamento
            cell.border = borda

    larguras = [12, 25, 18, 50, 15, 15, 15, 35, 35, 35]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 30

    wb.save(caminho)
    wb.close()

# ===============================================================
# SCRIPT PRINCIPAL — separação preservando CORES
# ===============================================================

BASE_DIR = Path(__file__).resolve().parent   # pasta do main.py
pasta_origem = BASE_DIR / "data"            # entrada: ./data
pasta_saida  = BASE_DIR / "output"          # saída: ./output

pasta_origem.mkdir(exist_ok=True)
pasta_saida.mkdir(exist_ok=True)

# grupos = { "SED AMAMBAI": [ {"linha": [...células...], "data": datetime}, ... ] }
grupos = {}

print("Lendo arquivos da pasta de origem...\n")

for arquivo in os.listdir(str(pasta_origem)):

    if not arquivo.lower().endswith((".xlsx", ".xlsm")):
        continue

    caminho = pasta_origem / arquivo   # <-- define o caminho do arquivo de entrada
    wb = load_workbook(str(caminho), data_only=True)
    ws = wb.active

    # identificar a coluna CLIENTE e DATA ULTIMA POSICAO
    col_cliente = None
    col_data = None

    for col in range(1, ws.max_column + 1):
        valor_cab = ws.cell(row=1, column=col).value
        if valor_cab is None:
            continue

        nome_normalizado = normalizar(valor_cab)

        if nome_normalizado == "CLIENTE":
            col_cliente = col
        if nome_normalizado == "DATA ULTIMA POSICAO":
            col_data = col

    if not col_cliente or not col_data:
        print(f"Erro no arquivo {arquivo}: coluna CLIENTE ou DATA ULTIMA POSICAO não encontrada.")
        wb.close()
        continue

    # processar linhas
    for r in range(2, ws.max_row + 1):
        cel_cliente = ws.cell(r, col_cliente).value

        if cel_cliente is None:
            continue

        cliente_raw = str(cel_cliente).strip()
        if not cliente_raw:
            continue

        # pega texto do cliente até o primeiro " - "
        cliente_base = cliente_raw.split(" - ")[0].strip()

        # copia linha inteira (10 primeiras colunas)
        linha = [ws.cell(r, c) for c in range(1, 11)]

        # pega a data da coluna correta
        cel_data_val = ws.cell(r, col_data).value
        data_linha = None

        if isinstance(cel_data_val, datetime):
            data_linha = cel_data_val
        elif isinstance(cel_data_val, date):
            # converte date para datetime (meia-noite)
            data_linha = datetime.combine(cel_data_val, datetime.min.time())
        elif isinstance(cel_data_val, str):
            try:
                data_linha = datetime.strptime(cel_data_val.strip(), "%d/%m/%Y")
            except ValueError:
                data_linha = None

        # salva no grupo
        if cliente_base not in grupos:
            grupos[cliente_base] = []

        grupos[cliente_base].append({
            "linha": linha,
            "data": data_linha
        })

    wb.close()

print("Separando arquivos por cliente...\n")

# ===============================================================
# Criar arquivos por cliente
# ===============================================================

for cliente, registros in grupos.items():

    # pegar a MAIOR DATA dentre os registros
    datas_validas = [reg["data"] for reg in registros if isinstance(reg["data"], datetime)]

    if datas_validas:
        data_final = max(datas_validas)
        sufixo_data = f" ({data_final.strftime('%d-%m-%Y')})"
    else:
        data_final = None
        sufixo_data = " (SEM-DATA)"

    # nome final do arquivo -> CLIENTE (dd-mm-aaaa).xlsx ou CLIENTE (SEM-DATA).xlsx
    nome_arq = f"{cliente}{sufixo_data}.xlsx"

    # higienizar caracteres proibidos em nome de arquivo no Windows
    for ch in '\\/:*?"<>|':
        nome_arq = nome_arq.replace(ch, "_")

    caminho_saida = pasta_saida / nome_arq

    # criar workbook
    novo = Workbook()
    ws_novo = novo.active

    # copiar cabeçalho do PRIMEIRO registro desse grupo
    primeira_linha = registros[0]["linha"]
    ws_original = primeira_linha[0].parent  # worksheet de origem da linha

    for c in range(1, 11):
        copiar_celula(ws_original.cell(row=1, column=c), ws_novo.cell(row=1, column=c))

    # copiar linhas preservando cores/estilos
    linha_dest = 2
    for reg in registros:
        linha = reg["linha"]
        for col_index, cel in enumerate(linha, start=1):
            copiar_celula(cel, ws_novo.cell(row=linha_dest, column=col_index))
        linha_dest += 1

    # salvar e formatar
    novo.save(str(caminho_saida))
    novo.close()
    formatar_arquivo_excel(str(caminho_saida))

    print(f"Arquivo criado: {caminho_saida}")

print("\nProcesso finalizado com sucesso!")
