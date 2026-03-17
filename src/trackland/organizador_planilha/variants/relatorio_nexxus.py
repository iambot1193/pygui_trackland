import unicodedata
from copy import copy
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ===============================================================
# FUNÇÃO: normalizar texto
# ===============================================================
def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = texto.replace("Ç", "C")
    return texto

# ===============================================================
# FUNÇÃO: copiar célula preservando formatação
# ===============================================================
def copiar_celula(origem, destino):
    destino.value = origem.value

    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)

# ===============================================================
# FUNÇÃO: formatar relatório
# ===============================================================
def formatar_arquivo_excel(caminho):
    wb = load_workbook(str(caminho))
    ws = wb.active

    fonte = Font(name="Calibri", size=12, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    fill_cabecalho = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    max_row = ws.max_row
    max_col = 10

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        if cell.fill.start_color.index in ("00000000", "FFFFFFFF"):
            cell.fill = fill_cabecalho
        cell.font = fonte
        cell.alignment = alinhamento
        cell.border = borda

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = fonte
            cell.alignment = alinhamento
            cell.border = borda

    larguras = [12, 25, 18, 50, 15, 15, 15, 35, 35, 35]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 30

    wb.save(str(caminho))
    wb.close()

# ===============================================================
# SCRIPT PRINCIPAL — NEXXUS
# ===============================================================

# Como este arquivo está em /variants, o root do projeto é a pasta acima:
ROOT_DIR = Path(__file__).resolve().parent.parent

pasta_origem = ROOT_DIR / "data"     # entrada: ../data
pasta_saida  = ROOT_DIR / "output"   # saída:  ../output

pasta_origem.mkdir(exist_ok=True)
pasta_saida.mkdir(exist_ok=True)

relatorio = Workbook()
ws_rel = relatorio.active

cabecalho_copiado = False
linha_dest = 2

print("Gerando relatório NEXXUS...\n")

for caminho in pasta_origem.iterdir():

    # só arquivos Excel
    if caminho.suffix.lower() not in (".xlsx", ".xlsm"):
        continue

    # ignorar temporários do Excel (~$)
    if caminho.name.startswith("~$"):
        continue

    wb = load_workbook(str(caminho), data_only=True)
    ws = wb.active

    # Localizar coluna CLIENTE
    col_cliente = None
    for col in range(1, ws.max_column + 1):
        nome = normalizar(ws.cell(1, col).value)
        if nome == "CLIENTE":
            col_cliente = col
            break

    if not col_cliente:
        print(f"Arquivo ignorado (sem coluna CLIENTE): {caminho.name}")
        wb.close()
        continue

    # Copiar cabeçalho apenas 1 vez
    if not cabecalho_copiado:
        for c in range(1, 11):
            copiar_celula(ws.cell(1, c), ws_rel.cell(1, c))
        cabecalho_copiado = True

    # Processar linhas
    for r in range(2, ws.max_row + 1):
        valor_cliente = ws.cell(r, col_cliente).value
        if not valor_cliente:
            continue

        cliente = str(valor_cliente).upper()

        # SOMENTE NEXXUS / NEXUS / NEX
        if "NEX" not in cliente:
            continue

        for c in range(1, 11):
            copiar_celula(ws.cell(r, c), ws_rel.cell(linha_dest, c))

        linha_dest += 1

    wb.close()

# Salvar relatório
caminho_relatorio = pasta_saida / "RELATORIO_NEXXUS.xlsx"
relatorio.save(str(caminho_relatorio))
relatorio.close()

formatar_arquivo_excel(caminho_relatorio)

print(f"\nRelatório criado: {caminho_relatorio}")
print("Processo concluído!")