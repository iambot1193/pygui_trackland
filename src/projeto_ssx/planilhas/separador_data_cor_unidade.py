import os
import unicodedata
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date

# ===============================================================
# FUNÇÃO: normalizar texto
# ===============================================================
def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = texto.replace("Ç", "C")
    return texto

# ===============================================================
# FUNÇÃO: copiar célula preservando formatação
# ===============================================================
def copiar_celula(origem, destino):
    destino.value = origem.value

    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)

# ===============================================================
# FUNÇÃO: formatar relatório
# ===============================================================
def formatar_arquivo_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    fonte = Font(name="Calibri", size=12, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    fill_cabecalho = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    max_row = ws.max_row
    max_col = 10

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        if cell.fill.start_color.index in ("00000000", "FFFFFFFF"):
            cell.fill = fill_cabecalho
        cell.font = fonte
        cell.alignment = alinhamento
        cell.border = borda

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = fonte
            cell.alignment = alinhamento
            cell.border = borda

    larguras = [12, 25, 18, 50, 15, 15, 15, 35, 35, 35]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 30

    wb.save(caminho)
    wb.close()

# ===============================================================
# SCRIPT PRINCIPAL — NEXXUS
# ===============================================================

pasta_origem = r"C:\Users\jenni\Downloads\SEDOBTER"
pasta_saida = r"C:\Users\jenni\Downloads\SSXSAIDA"
os.makedirs(pasta_saida, exist_ok=True)

relatorio = Workbook()
ws_rel = relatorio.active

cabecalho_copiado = False
linha_dest = 2

print("Gerando relatório NEXXUS...\n")

for arquivo in os.listdir(pasta_origem):

    # IGNORAR arquivos temporários do Excel (~$)
    if arquivo.startswith("~$"):
        continue

    if not arquivo.lower().endswith((".xlsx", ".xlsm")):
        continue

    caminho = os.path.join(pasta_origem, arquivo)
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    # Localizar coluna CLIENTE
    col_cliente = None
    for col in range(1, ws.max_column + 1):
        nome = normalizar(ws.cell(1, col).value)
        if nome == "CLIENTE":
            col_cliente = col
            break

    if not col_cliente:
        print(f"Arquivo ignorado (sem coluna CLIENTE): {arquivo}")
        wb.close()
        continue

    # Copiar cabeçalho apenas 1 vez
    if not cabecalho_copiado:
        for c in range(1, 11):
            copiar_celula(ws.cell(1, c), ws_rel.cell(1, c))
        cabecalho_copiado = True

    # Processar linhas
    for r in range(2, ws.max_row + 1):
        valor_cliente = ws.cell(r, col_cliente).value
        if not valor_cliente:
            continue

        cliente = str(valor_cliente).upper()

        # SOMENTE NEXXUS / NEXUS / NEX
        if "NEX" not in cliente:
            continue

        linha = [ws.cell(r, c) for c in range(1, 11)]
        for col_index, cel in enumerate(linha, start=1):
            copiar_celula(cel, ws_rel.cell(linha_dest, col_index))

        linha_dest += 1

    wb.close()

# Salvar relatório
caminho_relatorio = os.path.join(pasta_saida, "RELATORIO_NEXXUS.xlsx")
relatorio.save(caminho_relatorio)
relatorio.close()

formatar_arquivo_excel(caminho_relatorio)

print(f"\nRelatório criado: {caminho_relatorio}")
print("Processo concluído!")
